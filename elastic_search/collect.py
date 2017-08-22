from __future__ import print_function, unicode_literals
import traceback
from openpyxl import load_workbook
import os
import csv
from elastic_search.es_core_config import get_hash, create_connection
from elastic_search.es_models import DataHead
from elastic_search.es_settings import data_files, error_file, key_mapping, DATA_PATH, INDEX_NAME


def get_path(file_name):
    return os.path.join(DATA_PATH, file_name)


def log_error(entry):
    filename = os.path.join(DATA_PATH, error_file)

    if os.path.exists(filename):
        append_write = 'a'
    else:
        append_write = 'w'

    with open(filename, append_write) as f:
        f.write(entry)
        f.write(str("\n"))


# [FOR XLSX]
# def extract(file_name):
#     print("Loading File...")
#     wb = load_workbook(get_path(file_name), read_only=True)
#     print("File Loaded.")
#
#     ws = wb.get_sheet_by_name('Report 1')
#
#     headers = ws[4]
#     row_count = ws.max_row
#
#     print("Mining Started...")
#
#     for row_num, row in enumerate(ws.iter_rows(row_offset=4)):
#         ent = []
#         bpnum = "NA"
#         for col, cell in enumerate(row):
#             if str(headers[col].value).strip(" ") == "Bpartner":
#                 bpnum = str(cell.value)
#             if cell.value:
#                 ent.append(str(cell.value))
#
#         obj = DataHead(
#             bpnum=bpnum,
#             address="Not available yet",
#             keywords=ent
#         )
#         print(str((row_num / row_count) * 100.00) + '%' + ' Loaded\t[%s/%s]' % (row_num, row_count), end='\r')
#         yield obj


def get_row_count(file_name):
    return len(open(get_path(file_name)).readlines()) - 1


# [FOR CSV]
def extract_csv(file_name):
    print("Loading File...")
    row_count = get_row_count(file_name)
    fl_row_count = float(row_count)
    f = open(get_path(file_name), 'rb')
    reader = csv.DictReader(f)
    print("File Loaded.")

    print("Mining Started...")
    error_count = 0

    for row_num, row in enumerate(reader):
        try:
            for key in row.keys():
                row[key_mapping[key]] = row.pop(key).decode('utf-8', 'ignore').encode('ascii', 'ignore').decode('ascii')
            obj = DataHead(inhash=get_hash(row), **row)
        except Exception as e:
            log_error(str(row) + "[" + e.message + "][" + traceback.format_exc() + "]")
            error_count += 1
        else:
            print("%.2f" % ((row_num / fl_row_count) * 100.00) + ' %' + ' Loaded\t[%s/%s] ' % (row_num, row_count) +
                  '[Faults: %s]' % error_count, end='\r')
            yield obj


def collect_data():
    n = len(data_files)
    for foo_index, file_name in enumerate(data_files):
        index = foo_index + 1
        print("------ Part " + str(index) + " of " + str(n) + " started -------------")
        DataHead.bulk_create(extract_csv(file_name))
        print("------ Part " + str(index) + " of " + str(n) + " complete -------------")


def get_res_count(es, row):

    body = {
        "query": {
            "constant_score": {
                "filter": {
                    "term": {
                        "inhash": get_hash(row)
                    }
                }
            }
        }
    }
    res = es.search(
        index=INDEX_NAME,
        body=body,
        size=50
    )
    return res['hits']['total']


def read_csv(file_name):
    print("Loading File...")
    row_count = get_row_count(file_name)
    fl_row_count = float(row_count)
    f = open(get_path(file_name), 'rb')
    reader = csv.DictReader(f)
    print("File Loaded.")

    print("Checking Started...")
    es = create_connection()
    error_count = 0

    for row_num, row in enumerate(reader):
        try:
            for key in row.keys():
                row[key_mapping[key]] = row.pop(key).decode('utf-8', 'ignore').encode('ascii', 'ignore').decode('ascii')

            if get_res_count(es, row) == 0:
                obj = DataHead(inhash=get_hash(row), **row)
                obj.save()
                print("--------------------------------------------")
                print(row)
                print("---------New Entry Created------------------")
        except Exception as e:
            log_error(str(row) + "[" + e.message + "][" + traceback.format_exc() + "]")
            error_count += 1
        else:
            print("%.2f" % ((row_num / fl_row_count) * 100.00) + ' %' + ' Loaded\t[%s/%s]' % (row_num, row_count) +
                  '[Faults: %s]' % error_count, end='\r')


def check_data():
    print("---------Checking Indexes----------")
    n = len(data_files)
    for foo_index, file_name in enumerate(data_files):
        index = foo_index + 1
        print("------ Part " + str(index) + " of " + str(n) + " started -------------")
        read_csv(file_name)
        print("------ Part " + str(index) + " of " + str(n) + " complete -------------")


def validate_es_data():
    es = create_connection()
    body = {
        "query": {
            "match_all": {
            }
        }
    }

    res = es.search(
        index=INDEX_NAME,
        body=body,
        size=50
    )

    es_totals = res['hits']['total']

    n = 0

    for data_file in data_files:
        n += get_row_count(data_file)

    print("--------------Checking for errors---------")
    print(str(n) + " | " + str(es_totals))

    if n != es_totals:
        return False
    else:
        return True
